import os
import zipfile

import arrow
import openpyxl as opwb
import pandas as pd
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from sqlalchemy import func
from starlette.responses import FileResponse

from core.models.report import DailyPlan, Employee, WeeklyPlan
from core.repositories._base import BaseRepository
from core.services import get_db
from core.utils import assets_dir, underline_date, slash_date, from_date, to_date, current_time
from core.validators.report import ReportValidator, WeeklyPlanValidator


class ReportRepository(BaseRepository):

    def __init__(self):
        self.all_style = self.normal_style

    async def download_daily_report(self, id: int):
        db_client = get_db()
        db = next(db_client)
        date_with_underline = underline_date()
        date_with_slash = slash_date()

        res = (db.query(Employee, DailyPlan).filter(Employee.id == DailyPlan.user_id, Employee.id == id,
                                                    func.date(DailyPlan.create_time) >= from_date(),
                                                    func.date(DailyPlan.create_time) <= to_date()).first())

        name, plan, advice, tomorrow_plan = self.handle_res_from_mysql(res)

        # with open(self.get_daily_template(), 'r', encoding='utf-8') as daily_template:
        file_name = date_with_underline + f"_{name}.xlsx"
        file_path = os.path.join(self.get_save_daily_dir(), file_name)

        res = self.get_daily_template(name, date_with_slash, plan, advice, tomorrow_plan)
        df = pd.DataFrame(res)
        df.columns = self.columns
        df.to_excel(file_path, index=False)
        wb = opwb.load_workbook(file_path)
        ws = wb.worksheets[0]
        self.alter_daily_style(ws, len(plan))
        wb.save(file_path)
        file_name = os.path.basename(file_path)

        return FileResponse(file_path, filename=file_name)

    async def download_daily_report_all(self):
        date_with_underline = underline_date()
        date_with_slash = slash_date()
        file_name = date_with_underline + f"_日报.xlsx"
        file_path = os.path.join(self.get_save_daily_dir(), file_name)

        db_client = get_db()
        db = next(db_client)
        results = (db.query(Employee, DailyPlan)
                   .filter(Employee.id == DailyPlan.user_id,
                           func.date(DailyPlan.create_time) >= from_date(),
                           func.date(DailyPlan.create_time) <= to_date()).all())

        df_datas = []
        plans = []

        for res in results:
            name, plan, advice, tomorrow_plan = self.handle_res_from_mysql(res)
            res = self.get_daily_template(name, date_with_slash, plan, advice, tomorrow_plan)
            df = pd.DataFrame(res)
            df.columns = self.columns
            df_datas.append((name, df))
            plans.append(plan)

        with pd.ExcelWriter(file_path) as writer:
            for name, df in df_datas:
                df.to_excel(writer, sheet_name=name, index=False)

        wb = opwb.load_workbook(file_path)
        for ws, plan in zip(wb.worksheets, plans):
            self.alter_daily_style(ws, len(plan))

        wb.save(file_path)
        file_name = os.path.basename(file_path)
        return FileResponse(file_path, filename=file_name)

    async def download_daily_report_weekly(self):
        date_with_underline = underline_date()
        date_with_slash = slash_date()
        file_path = os.path.join(self.get_save_daily_dir(), date_with_underline + '_weekly')

        if not os.path.exists(file_path):
            os.mkdir(file_path)

        for i in range(20):
            file_name = date_with_underline + f'_{i}' + ".xlsx"
            # with open(self.get_daily_template(), 'r', encoding='utf-8') as daily_template:
            file_path_ = os.path.join(file_path, file_name)

            columns = ['工作计划表（日报）'] * 6
            # df = pd.read_excel(self.get_daily_template())
            name = '孟泽儒'
            advice = '111'
            tomorrow_plan = '222'
            plan = [{
                '1': 1,
                '2': '',
                '3': '',
                '4': '',
                '5': '',
                '6': ''
            }, {
                '1': 2,
                '2': '',
                '3': '',
                '4': '',
                '5': '',
                '6': ''
            }, {
                '1': 2,
                '2': '',
                '3': '',
                '4': '',
                '5': '',
                '6': ''
            }, {
                '1': 2,
                '2': '',
                '3': '',
                '4': '',
                '5': '',
                '6': ''
            }]
            res = self.get_daily_template(name, date_with_slash, plan, advice, tomorrow_plan)
            df = pd.DataFrame(res)
            df.columns = columns
            # df.fillna(method='ffill')
            # print(df)
            # df.iloc
            # df.iloc[0, 1:2] = 'mzr'

            df.to_excel(file_path_, index=False)
            wb = opwb.load_workbook(file_path_)
            ws = wb.worksheets[0]
            self.alter_daily_style(ws, len(plan))
            wb.save(file_path_)
            # df = pd.DataFrame(result)
            # df.columns = ["序号", "姓名", "年龄"]
            # ['存在问题及建议','明天工作计划']
            #
            # df.to_excel(file_path, index=False)
            # return FileResponse(file, filename="user.xlsx")
        zip_file_path = file_path + '.zip'
        file_name = os.path.basename(zip_file_path)
        # with tarfile.open(zip_file_path, "w") as tar:
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as f:
            os.chdir(file_path)
            for _, _, files in os.walk(file_path):
                # print(file)
                for file in files:
                    f.write(file)
            # f.write(file_path)
        return FileResponse(zip_file_path, filename=file_name)

    def add_daily_report(self, payload: ReportValidator):
        db_client = get_db()
        db = next(db_client)
        user_id = payload.user_id
        db.query(Employee).filter(Employee.id == payload.user_id).update({'today_condition': 1})

        payload = payload.dict()
        payload['create_time'] = current_time()

        new_daily_plan = DailyPlan(**payload)
        db.add(new_daily_plan)
        db.flush()

        return {'user_id': user_id}

    @staticmethod
    def get_save_daily_dir():
        return os.path.join(assets_dir(), 'excel', 'daily')

    def alter_daily_style(self, worksheet, plan_length):
        n_style = self.all_style
        for i in worksheet.rows:
            for j in i:
                j.style = n_style
        cell_title = worksheet['A1']
        cell_title.font = self.font_big
        cell_title.alignment = self.center_alignment
        cell_plan = worksheet['A3']
        cell_plan.font = self.font_mid
        cell_plan.alignment = self.center_alignment
        worksheet.merge_cells('A1:F1')  # 工作计划表（日报）
        worksheet.merge_cells('A3:F3')  # 工作计划
        worksheet.merge_cells('B2:C2')  # 记录人
        worksheet.merge_cells('E2:F2')  # 日期
        worksheet.merge_cells('B4:C4')  # 内容
        worksheet.merge_cells('D4:E4')  # 完成情况

        # 合并plan
        for i in range(5, 8):
            worksheet.merge_cells(f'B{i}:C{i}')
            worksheet.merge_cells(f'D{i}:E{i}')
        start_index = 8
        end_index = plan_length + start_index - 3
        if plan_length > 3:
            for index in range(start_index, end_index):
                worksheet.merge_cells(f'B{index}:C{index}')
                worksheet.merge_cells(f'D{index}:E{index}')

        # 存在问题及建议
        worksheet.merge_cells(f'A{end_index}:C{end_index}')
        worksheet.merge_cells(f'D{end_index}:F{end_index}')
        # 明天工作计划
        worksheet.merge_cells(f'A{end_index + 1}:C{end_index + 1}')
        worksheet.merge_cells(f'D{end_index + 1}:F{end_index + 1}')
        return

    def get_daily_template(self, name, date_with_slash, plan, advice, tomorrow_plan):
        res = self.title_template(name, date_with_slash)

        if len(plan) <= 3:
            plan += [{
                '1': len(plan) + 1,
                '2': '',
                '3': '',
                '4': '',
                '5': '',
                '6': ''
            }] * (3 - len(plan))
        tomorrow_plan_title = self.tomorrow_plan_title_template(advice, tomorrow_plan)
        res += plan
        res += tomorrow_plan_title
        return res

    @property
    def normal_style(self):
        return NamedStyle(
            name=f'all',
            font=Font(size=14),
            border=Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin')))

    @property
    def font_big(self):
        return Font(size=22)

    @property
    def font_mid(self):
        return Font(size=16, bold=False)

    @property
    def center_alignment(self):
        return Alignment(horizontal='center')

    @property
    def columns(self):
        return ['工作计划表（日报）'] * 6

    @staticmethod
    def title_template(name, date_with_slash):
        return [{
            '1': '记录人',
            '2': name,
            '3': name,
            '4': '日期',
            '5': date_with_slash,
            '6': date_with_slash
        }, {
            '1': '工作计划',
            '2': '工作计划',
            '3': '工作计划',
            '4': '工作计划',
            '5': '工作计划',
            '6': '工作计划'
        }, {
            '1': '序号',
            '2': '内容',
            '3': '内容',
            '4': '完成情况',
            '5': '完成情况',
            '6': '未完成原因'
        }]

    @staticmethod
    def tomorrow_plan_title_template(advice, tomorrow_plan):
        return [
            {
                '1': '存在问题及建议',
                '2': '存在问题及建议',
                '3': '存在问题及建议',
                '4': '明天工作计划',
                '5': '明天工作计划',
                '6': '明天工作计划'
            }, {
                '1': advice,
                '2': advice,
                '3': advice,
                '4': tomorrow_plan,
                '5': tomorrow_plan,
                '6': tomorrow_plan
            }
        ]

    def handle_res_from_mysql(self, res):
        plan = [{'1': 1, '2': '', '3': '', '4': '', '5': '', '6': ''},
                {'1': 2, '2': '', '3': '', '4': '', '5': '', '6': ''},
                {'1': 3, '2': '', '3': '', '4': '', '5': '', '6': ''}]
        if res:
            name = res[0].name
            res: DailyPlan = res[1]
            plan_ = res.daily_plan if res.daily_plan is not None else []
            advice = res.advice if res.advice is not None else ''
            tomorrow_plan = res.tomorrow_plan if res.tomorrow_plan is not None else ''
        else:
            name = '没有该条记录'
            plan_ = []
            advice = ''
            tomorrow_plan = ''
        for i, p in enumerate(plan_):
            if i + 1 > 3:
                plan.append(
                    {'1': i + 1, '2': p['content'], '3': p['content'], '4': p['condition'], '5': p['condition'],
                     '6': p['cause']})
                continue
            plan[i]['2'] = p['content']
            plan[i]['3'] = p['content']
            plan[i]['4'] = p['condition']
            plan[i]['5'] = p['condition']
            plan[i]['6'] = p['cause']
        return name, plan, advice, tomorrow_plan

    def get_daily_report_by_time(self, user_id: int, time: str):
        db_client = get_db()
        db = next(db_client)

        from_time = arrow.get(time).format('YYYY-MM-DD HH:mm:ss')
        to_time = arrow.get(time).shift(hours=23, minutes=59, seconds=59).format('YYYY-MM-DD HH:mm:ss')

        res = db.query(DailyPlan).filter(DailyPlan.user_id == user_id,
                                         func.date(DailyPlan.create_time) >= from_time,
                                         func.date(DailyPlan.create_time) <= to_time).first()
        if not res:
            return {'daily_plan': [],
                    'advice': '',
                    'tomorrow_plan': ''}
        return res.to_dict()

    def get_nearly_seven_days_daily_report_edit_time(self, id: int):
        db_client = get_db()
        db = next(db_client)
        current_day_time = to_date()
        seven_days_before_time = arrow.get(from_date()).shift(days=-6).format('YYYY-MM-DD HH:mm:ss')
        results = db.query(DailyPlan).filter(DailyPlan.user_id == id,
                                             func.date(DailyPlan.create_time) >= seven_days_before_time,
                                             func.date(DailyPlan.create_time) <= current_day_time).all()
        current_day_time_day = current_day_time.split(' ')[0]
        seven_days_before_time_day = seven_days_before_time.split(' ')[0]

        delta = arrow.get(current_day_time_day) - arrow.get(seven_days_before_time_day)

        all_7_days = {}

        for i in range(delta.days + 1):
            d = arrow.get(seven_days_before_time_day).shift(days=i).format('MM-DD')
            all_7_days[d] = {'day': d, 'time': ''}

        for r in results:
            create_time = str(r.create_time)
            day, time = create_time.split(' ')[0], create_time
            day = day.split('-', 1)[1]
            all_7_days[day]['time'] = time

        res = list(map(lambda x: x[1], all_7_days.items()))

        return res

    def add_or_update_weekly_plan(self, payload: WeeklyPlanValidator):
        db_client = get_db()
        db = next(db_client)
        p = payload.dict()
        is_update = p.pop('update')
        if is_update is False:
            p['create_time'] = current_time()

            wp = WeeklyPlan(**p)
            db.add(wp)
            db.flush()
            id = wp.id
            p.update({'id': id})
            res = p
        else:
            id = p.pop('id')

            db.query(WeeklyPlan).filter(WeeklyPlan.id == id).update(p)
            db.flush()
            res = p.update({'id': id})

        return res

    # def update_weekly_plan(self, id: int, payload: WeeklyPlanValidator):
    #     db_client = get_db()
    #     db = next(db_client)
    #
    #     res = db.query(WeeklyPlan).filter(WeeklyPlan.id == id).update(**payload.dict())
    #     db.flush()
    #     print(res)
    #     return res

    def get_weekly_plan(self, id: int):
        db_client = get_db()
        db = next(db_client)

        res = db.query(WeeklyPlan).filter(WeeklyPlan.user_id == id).all()
        results = []
        for r in res:
            r = r.to_dict()
            end_time = arrow.get(r['end_time']).format('YYYY-MM-DD')
            start_time = arrow.get(r['start_time']).format('YYYY-MM-DD')
            create_time = arrow.get(r['create_time']).format('YYYY-MM-DD')
            r.update({'end_time': end_time, 'start_time': start_time, 'create_time': create_time})
            results.append(r)
        return results
